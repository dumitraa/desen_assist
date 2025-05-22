import shutil
from pathlib import Path
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QPushButton, QProgressBar, QMessageBox, QLineEdit, QLabel # type: ignore
from qgis.core import QgsProject, QgsMessageLog, Qgis # type: ignore
import pandas as pd
from openpyxl import load_workbook
import os
from .. import conf

class GenerateExcelDialog(QDialog):
    def __init__(self, base_dir):
        super().__init__()
        self.base_dir = base_dir
        self.setWindowTitle("Generare Excel Strazi Lipsa")
        self.layout = QVBoxLayout()

        # Add input for locality
        self.locality_label = QLabel("Cod_Localitate:")
        self.layout.addWidget(self.locality_label)
        
        self.locality_input = QLineEdit(self)
        self.layout.addWidget(self.locality_input)

        self.progress_bar = QProgressBar(self)
        self.layout.addWidget(self.progress_bar)

        self.run_button = QPushButton("Genereaza Excel", self)
        self.run_button.clicked.connect(self.__exec__)
        self.layout.addWidget(self.run_button)

        self.setLayout(self.layout)
    
    @staticmethod
    def plugin_path(*args) -> Path:
        """ Return the path to the plugin root folder or file. """
        path = Path(__file__).resolve().parent
        for item in args:
            path = path.joinpath(item)
        return path
    
    def __exec__(self):
        try:
            self.progress_bar.setValue(10)
            layers = {"STALP_JT": None, "BRANS_FIRI_GRPM_JT": None}
            for layer_name in layers.keys():
                found_layers = QgsProject.instance().mapLayersByName(layer_name)
                if not found_layers:
                    raise ValueError(f"Layer {layer_name} not found!")
                layers[layer_name] = found_layers[0]
            
            self.progress_bar.setValue(30)
            nomenclator_path = self.plugin_path('templates', 'nomenclator.xlsx')
            if not nomenclator_path.exists():
                raise FileNotFoundError("nomenclator.xlsx not found!")
            nomenclator = pd.ExcelFile(nomenclator_path)

            locality = self.locality_input.text().strip()
            if not locality:
                self.show_message("Introdu o localitate!", error=True)
                return
            
            COD_LOC = 'COD_LOC'
            STR = 'NUME_STR'
            TIP_STR = 'TIP_STR'
            
            
            self.progress_bar.setValue(50)
            df = nomenclator.parse('Sheet1', dtype=str)
            QgsMessageLog.logMessage(f"Localitati: {df[COD_LOC]}", "DesenAssist", level=Qgis.Info)
            city_row = df[df[COD_LOC] == locality]
            if city_row.empty:
                raise ValueError(f"Localitatea {locality} nu a fost gasita in nomenclator!")
            city_code = city_row.iloc[0][COD_LOC]
            
            known_streets = set(zip(df[df[COD_LOC] == city_code][STR], df[df[COD_LOC] == city_code][TIP_STR]))

            
            layer_streets = set()
            for layer in layers.values():
                for feature in layer.getFeatures():
                    street_name = str(feature["STR"])
                    street_type = str(feature["TIP_STR"])
                    if street_name and street_type:
                        layer_streets.add((street_name, street_type))
            
            self.progress_bar.setValue(70)
            missing_streets = sorted(layer_streets - known_streets)
            QgsMessageLog.logMessage(f"Missing streets: {missing_streets}", "DesenAssist", level=Qgis.Info)
            
            if missing_streets:
                self.write_missing_streets_to_excel(missing_streets, city_row)
                self.progress_bar.setValue(100)
                self.show_message(f"File generation completed successfully! {len(missing_streets)} streets missing. Path: {self.base_dir}", error=False)
            else:
                self.progress_bar.setValue(100)
                self.show_message(f"Toate strazile din {locality} sunt in nomenclator!", error=False)
            
        except Exception as e:
            QgsMessageLog.logMessage(f"Error during execution: {e}", "DesenAssist", level=Qgis.Critical)
            self.show_message(f"Error: {e}", error=True)
            
    def create_valid_output(self, main_dir, filename, subdir=None):
        if subdir:
            full_path = os.path.join(main_dir, subdir)
        else:
            full_path = main_dir

        # Ensure directory exists
        os.makedirs(full_path, exist_ok=True)

        # Normalize path to avoid mix of \ and /
        valid_path = os.path.normpath(os.path.join(full_path, filename)).replace("\\", "/")
        
        return valid_path

    
    def write_missing_streets_to_excel(self, missing_streets, city_row):
        new_file_name = f"Tabel_completare strazi in nomenclatorul de adrese_{self.locality_input.text().strip()}.xlsx"
        output_file = self.create_valid_output(self.base_dir, new_file_name)
        template_path = self.plugin_path('templates', 'to_complete.xlsx')

        # Copy template to output location
        try:
            shutil.copyfile(template_path, output_file)
        except Exception as e:
            QgsMessageLog.logMessage(f"Error copying template: {e}", "DesenAssist", level=Qgis.Critical)
            return

        try:
            workbook = load_workbook(output_file)
            sheet = workbook["Sheet1"]
        except Exception as e:
            QgsMessageLog.logMessage(f"Error loading workbook: {e}", "DesenAssist", level=Qgis.Critical)
            return

        start_row = 2
        header_row = 1
        
        existing_headers = {sheet.cell(row=header_row, column=col_idx).value: col_idx for col_idx in range(1, sheet.max_column + 1) if sheet.cell(row=header_row, column=col_idx).value}

        if not existing_headers:
            QgsMessageLog.logMessage("No headers found in template!", "DesenAssist", level=Qgis.Critical)
            return

        headers = ["Judet", "Cod_Comuna(UAT)", "Nume_Comuna(UAT)", "Cod_Localitate", "Nume_Localitate", "Nume Strada", "Tip / STRTYPEAB", "CP / POST_CODE", "GrpStrReg / REGIOGROUP", "REGPOLIT"]

        for row_idx, (street, tip_strada) in enumerate(missing_streets, start=start_row):
            row_data = [
                str(city_row.iloc[0]['JUDET']) if str(city_row.iloc[0]['JUDET']) not in conf.NULL_VALUES else "",
                str(city_row.iloc[0]['COD_UAT']) if str(city_row.iloc[0]['COD_UAT']) not in conf.NULL_VALUES else "",
                str(city_row.iloc[0]['NUME_UAT']) if str(city_row.iloc[0]['NUME_UAT']) not in conf.NULL_VALUES else "",
                str(city_row.iloc[0]['COD_LOC']) if str(city_row.iloc[0]['COD_LOC']) not in conf.NULL_VALUES else "",
                str(city_row.iloc[0]['NUME_LOC']) if str(city_row.iloc[0]['NUME_LOC']) not in conf.NULL_VALUES else "",
                str(street),
                str(tip_strada),
                str(city_row.iloc[0]['POST_CODE']) if str(city_row.iloc[0]['POST_CODE']) not in conf.NULL_VALUES else "",
                str(city_row.iloc[0]['REGIOGROUP']) if str(city_row.iloc[0]['REGIOGROUP']) not in conf.NULL_VALUES else "",
                str(city_row.iloc[0]['REGPOLIT']) if str(city_row.iloc[0]['REGPOLIT']) not in conf.NULL_VALUES else ""
            ]
            
            QgsMessageLog.logMessage(f"Row data: {row_data}", "DesenAssist", level=Qgis.Info)

            for col_idx, (header, cell_value) in enumerate(zip(headers, row_data), start=1):
                if str(header).strip() in existing_headers:
                    cell = sheet.cell(row=row_idx, column=existing_headers[header], value=cell_value if cell_value is not None else "")
                    cell.number_format = "@"

        try:
            workbook.save(output_file)
        except Exception as e:
            QgsMessageLog.logMessage(f"Error saving workbook: {e}", "DesenAssist", level=Qgis.Critical)

    
    def show_message(self, message, error=False):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Critical if error else QMessageBox.Information)
        msg_box.setWindowTitle("Error" if error else "Complete")
        msg_box.setText(message)
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec_()
        self.close()